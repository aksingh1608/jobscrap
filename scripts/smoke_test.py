"""Offline checks for normalize → dedup → filter → rank → store → export.

Runs without network access and without the embedding model, so CI can use it
as a fast gate before the real scrape. Exits non-zero on the first failure.
"""

from __future__ import annotations

import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.config import load_config
from src.dedup import dedup_jobs
from src.export import export_excel, prune_exports
from src.filter import filter_jobs
from src.models import JobRecord
from src.normalize import canonical_url, clean_title, normalize_jobs
from src.rank import rank_jobs
from src.store import JobStore
from src.textmatch import any_match, contains


def _days_ago(n: int) -> str:
    return (date.today() - timedelta(days=n)).isoformat()


SAMPLES = [
    JobRecord(
        title="Machine Learning Internship",
        company="AI Labs Berlin",
        location="Berlin, Germany",
        source="test",
        url="https://example.com/jobs/ml-intern-1?utm_source=newsletter",
        description=(
            "We are looking for an Internship in Machine Learning. "
            "Working language: English. Python, PyTorch, NLP/LLM projects. "
            "German is a plus but not required."
        ),
        posted_date=_days_ago(1),
    ),
    JobRecord(
        title="Senior ML Engineer",
        company="BigCorp",
        location="Berlin",
        source="test",
        url="https://example.com/jobs/senior-ml",
        description="Permanent senior role. 5+ years experience. Lead the team.",
        posted_date=_days_ago(1),
    ),
    JobRecord(
        title="Werkstudent Data Science",
        company="DataCo",
        location="München, Deutschland",
        source="test",
        url="https://example.com/jobs/ws-ds",
        description=(
            "Werkstudent Data Science mit Python und Machine Learning. "
            "Deutschkenntnisse mindestens C1 erforderlich, verhandlungssicher."
        ),
        posted_date=_days_ago(2),
    ),
    JobRecord(
        title="HiWi Computer Vision",
        company="TU Berlin",
        location="Berlin",
        source="arbeitsagentur",
        url="https://example.com/jobs/hiwi-cv",
        description=(
            "Studentische Hilfskraft / HiWi for Computer Vision research. "
            "English working language. Deep Learning with PyTorch."
        ),
        posted_date=_days_ago(0),
    ),
    JobRecord(
        title="Machine Learning Internship - Apply now",
        company="AI Labs Berlin",
        location="Berlin",
        source="jobspy:indeed",
        url="https://www.example.com/jobs/ml-intern-1/",
        description="Duplicate of the first posting, reached via a different board.",
        posted_date=_days_ago(1),
    ),
    # The regression that motivated the rewrite: matched "ai" inside "e-mail"
    # and "data" inside marketing copy, and scored as high as a real ML role.
    JobRecord(
        title="E-commerce Intern",
        company="ShopCo",
        location="Berlin",
        source="jobspy:indeed",
        url="https://example.com/jobs/ecom-intern",
        description=(
            "Internship in e-commerce. Send your e-mail to apply. You will "
            "help with campaigns and available product listings. Great team!"
        ),
        posted_date=_days_ago(1),
    ),
]

FAILURES: list[str] = []


def check(condition: bool, message: str) -> None:
    if condition:
        print(f"  ok   {message}")
    else:
        print(f"  FAIL {message}")
        FAILURES.append(message)


def test_textmatch() -> None:
    print("textmatch")
    check(not contains("send your e-mail to apply", "ai"), "'ai' does not match 'e-mail'")
    check(not contains("available positions", "ai"), "'ai' does not match 'available'")
    check(contains("we use AI models", "ai"), "'ai' matches the standalone word")
    check(contains("machine-learning team", "machine learning"), "hyphens match spaces")
    check(any_match("Deep Learning role", ["nlp", "deep learning"]), "any_match finds a phrase")


def test_zero_is_a_real_setting() -> None:
    """`cfg.get(k) or default` swallowed 0, so --min-score 0 silently used 45."""
    print("numeric config")
    from src.config import num

    check(num({"min_fit_score": 0}, "min_fit_score", 55) == 0.0, "explicit 0 is kept")
    check(num({}, "min_fit_score", 55) == 55.0, "missing key falls back")
    check(num({"min_fit_score": None}, "min_fit_score", 55) == 55.0, "null falls back")
    check(num({"min_fit_score": "x"}, "min_fit_score", 55) == 55.0, "garbage falls back")

    sample = JobRecord(
        title="Data Science Internship",
        company="Co",
        location="Berlin",
        source="test",
        url="https://example.com/z",
        description="Machine learning internship with Python. English working language.",
        posted_date=_days_ago(0),
    )
    cfg = load_config()
    scored = rank_jobs([sample], {**cfg, "min_fit_score": 0, "min_results": 0})
    check(len(scored) == 1, "min_fit_score=0 keeps everything instead of defaulting")


def test_circuit_breaker() -> None:
    print("circuit breaker")
    from src.http_util import CircuitBreaker

    cb = CircuitBreaker("test", threshold=3)
    check(bool(cb), "breaker starts closed")
    cb.record_failure()
    cb.record_failure()
    check(bool(cb), "breaker stays closed below the threshold")
    cb.record_failure()
    check(not bool(cb), "breaker trips at the threshold")

    cb2 = CircuitBreaker("test2", threshold=3)
    cb2.record_failure()
    cb2.record_failure()
    cb2.record_success()
    cb2.record_failure()
    check(bool(cb2), "a success resets the consecutive failure count")


def test_normalize_helpers() -> None:
    print("normalize helpers")
    check(
        canonical_url("https://www.example.com/jobs/x/?utm_source=a&id=7")
        == "https://example.com/jobs/x?id=7",
        "canonical_url strips utm + www + trailing slash, keeps real params",
    )
    check(
        clean_title("  Data Science Intern  -  Apply Now ") == "Data Science Intern",
        "clean_title drops call-to-action suffixes",
    )


def test_pipeline_stages(cfg: dict) -> list[JobRecord]:
    print("pipeline stages")
    norm = normalize_jobs(list(SAMPLES), cfg)
    urls = [j.url for j in norm]
    check(len(urls) == len(set(urls)), "normalize leaves no duplicate canonical URLs")

    unique = dedup_jobs(norm)
    titles = [j.title for j in unique]
    check(
        sum(1 for t in titles if "Machine Learning Internship" in t) == 1,
        "cross-board duplicate collapses to one row",
    )

    filtered = filter_jobs(unique, cfg)
    kept = {j.title for j in filtered}
    check(not any("Senior" in t for t in kept), "senior role is filtered out")
    check(
        not any("E-commerce" in t for t in kept),
        "e-commerce internship no longer passes the AI/ML filters",
    )
    check(
        not any("Werkstudent Data Science" == t for t in kept),
        "C1-German posting is filtered out",
    )
    check(len(filtered) >= 1, "at least one genuine match survives")
    return filtered


def test_ranking(filtered: list[JobRecord], cfg: dict) -> list[JobRecord]:
    print("ranking")
    # min_fit_score 0 so we can inspect the full spread offline
    ranked = rank_jobs(filtered, {**cfg, "min_fit_score": 0})
    check(bool(ranked), "ranking returns scored jobs")
    check(
        all(0 <= j.fit_score <= 100 for j in ranked),
        "every fit score is inside 0–100",
    )
    check(
        all(j.fit_reason for j in ranked),
        "every job carries a fit reason",
    )
    check(
        ranked == sorted(ranked, key=lambda j: j.fit_score, reverse=True),
        "results come back sorted by fit score",
    )
    top = ranked[0]
    check(
        "Machine Learning" in top.title or "Computer Vision" in top.title,
        f"a real AI role ranks first (got {top.title!r} at {top.fit_score})",
    )
    return ranked


def test_migration_from_old_schema(tmp: Path) -> None:
    """Opening a pre-last_seen database must upgrade it, not crash."""
    print("schema migration")
    import sqlite3

    db = tmp / "legacy.db"
    conn = sqlite3.connect(db)
    conn.executescript(
        """
        CREATE TABLE jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL, company TEXT NOT NULL, location TEXT NOT NULL,
            source TEXT NOT NULL, url TEXT NOT NULL UNIQUE, description TEXT,
            language TEXT, role_type TEXT, posted_date TEXT, fit_score REAL,
            fit_reason TEXT, first_seen TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'new'
        );
        CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT);
        INSERT INTO jobs (title, company, location, source, url, first_seen, fit_score)
        VALUES ('Old Job', 'OldCo', 'Berlin', 'test', 'https://example.com/old',
                '2026-01-01', 70);
        """
    )
    conn.commit()
    conn.close()

    store = JobStore(str(db))
    check(True, "legacy database opens without error")
    rows = store.active_for_export()
    check(len(rows) == 1, "existing rows survive the migration")
    check(rows[0].last_seen == "2026-01-01", "last_seen backfills from first_seen")


def test_store_and_export(ranked: list[JobRecord], cfg: dict) -> None:
    print("store + export")
    store = JobStore(cfg["paths"]["sqlite"])
    store.upsert_ranked(ranked)
    active = store.active_for_export()
    check(len(active) == len(ranked), "all ranked jobs are stored as open")
    check(
        all(j.is_new_today() for j in active),
        "jobs stored today are flagged as new today",
    )

    # A job nobody has seen for a month should leave the sheet.
    with store._connect() as conn:  # noqa: SLF001 — test reaches in on purpose
        conn.execute(
            "UPDATE jobs SET last_seen=? WHERE url=?",
            (_days_ago(40), active[0].url),
        )
    check(store.expire_stale(21) == 1, "stale job is expired")
    check(len(store.active_for_export()) == len(active) - 1, "expired job leaves the sheet")

    path = export_excel(store.active_for_export(), cfg["paths"]["exports_dir"])
    check(path.exists() and path.stat().st_size > 0, "excel workbook is written")

    from openpyxl import load_workbook

    wb = load_workbook(path)
    check(wb.sheetnames == ["All Open", "New Today"], "workbook has both sheets")
    check(wb["All Open"].max_row >= 2, "workbook has data rows")

    check(prune_exports(cfg["paths"]["exports_dir"], keep=30) == 0, "prune keeps recent exports")


def main() -> int:
    cfg = load_config()
    with tempfile.TemporaryDirectory() as tmp:
        cfg["paths"]["sqlite"] = str(Path(tmp) / "smoke.db")
        cfg["paths"]["exports_dir"] = str(Path(tmp) / "exports")

        test_textmatch()
        test_zero_is_a_real_setting()
        test_circuit_breaker()
        test_normalize_helpers()
        filtered = test_pipeline_stages(cfg)
        ranked = test_ranking(filtered, cfg)
        test_migration_from_old_schema(Path(tmp))
        test_store_and_export(ranked, cfg)

    if FAILURES:
        print(f"\n{len(FAILURES)} check(s) FAILED:")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print("\nAll smoke checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
