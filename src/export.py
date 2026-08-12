"""Excel export: dated workbook with a New Today sheet and clickable apply links."""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models import JobRecord, today_iso

log = logging.getLogger(__name__)

COLUMNS = [
    ("New", 7),
    ("Fit Score", 10),
    ("Title", 44),
    ("Company", 26),
    ("Location", 22),
    ("Source", 20),
    ("Language", 10),
    ("Posted Date", 13),
    ("Age (days)", 11),
    ("Fit Reason", 62),
    ("Apply Link", 46),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
LINK_FONT = Font(color="0563C1", underline="single")

# Fit score bands → row tint. Makes the good rows findable at a glance.
SCORE_BANDS = [
    (80, PatternFill("solid", fgColor="C6EFCE")),  # strong
    (65, PatternFill("solid", fgColor="E2EFDA")),  # good
    (50, PatternFill("solid", fgColor="FFF7E6")),  # worth a look
]


def _age_days(job: JobRecord) -> Optional[int]:
    if not job.posted_date:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            posted = datetime.strptime(str(job.posted_date)[:10], fmt).date()
        except ValueError:
            continue
        return max(0, (date.today() - posted).days)
    return None


def _score_fill(score: float) -> Optional[PatternFill]:
    for threshold, fill in SCORE_BANDS:
        if score >= threshold:
            return fill
    return None


def _write_sheet(ws: Worksheet, jobs: list[JobRecord]) -> None:
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, job in enumerate(jobs, start=2):
        is_new = job.is_new_today()
        age = _age_days(job)
        values = [
            "NEW" if is_new else "",
            job.fit_score,
            job.title,
            job.company,
            job.location,
            job.source,
            job.language,
            job.posted_date or "",
            age if age is not None else "",
            job.fit_reason,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        link_cell = ws.cell(row=row_idx, column=len(COLUMNS), value=job.url)
        if job.url:
            link_cell.hyperlink = job.url
            link_cell.font = LINK_FONT

        fill = _score_fill(job.fit_score)
        if fill:
            for col_idx in range(2, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        if is_new:
            new_cell = ws.cell(row=row_idx, column=1)
            new_cell.fill = NEW_FILL
            new_cell.font = Font(bold=True, color="9C5700")

    last_row = max(2, len(jobs) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"


def export_excel(jobs: list[JobRecord], exports_dir: str) -> Path:
    """Write the dated workbook. Sheet 1 = everything open, sheet 2 = new today."""
    Path(exports_dir).mkdir(parents=True, exist_ok=True)
    out = Path(exports_dir) / f"internships_{today_iso()}.xlsx"

    ranked = sorted(jobs, key=lambda j: j.fit_score, reverse=True)
    new_today = [j for j in ranked if j.is_new_today()]

    wb = Workbook()
    ws = wb.active
    ws.title = "All Open"
    _write_sheet(ws, ranked)

    # Put the day's fresh finds on their own sheet so nothing gets buried.
    ws_new = wb.create_sheet("New Today")
    _write_sheet(ws_new, new_today)

    wb.save(out)
    log.info(
        "Excel exported: %s (%s rows, %s new today)", out, len(ranked), len(new_today)
    )
    return out


def prune_exports(exports_dir: str, keep: int) -> int:
    """Keep only the newest `keep` dated workbooks so the repo stays small."""
    if keep <= 0:
        return 0
    files = sorted(
        Path(exports_dir).glob("internships_*.xlsx"),
        key=lambda p: p.name,
        reverse=True,
    )
    removed = 0
    for stale in files[keep:]:
        try:
            stale.unlink()
            removed += 1
        except OSError:
            log.warning("Could not remove old export %s", stale, exc_info=True)
    if removed:
        log.info("Pruned %s old export file(s), keeping newest %s", removed, keep)
    return removed
